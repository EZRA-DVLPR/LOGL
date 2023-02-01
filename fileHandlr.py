#this file handles all functions for files
#including excel files and txt files

import webHandlr
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

#creates/opens the file to be written to
#returns the workbook
def wbChecker():

    # Try to open the workbook
    # else create it
    try:
        wb = load_workbook('gamelist_wb.xlsx')

        #check if Wishlist worksheet exists and if not then create worksheet named wishlist
        if (not (wb.sheetnames[1])):
            wb.create_sheet("Wishlist")

        #set active worksheet to owned games
        wb.active = wb['Owned Games']

        print("Excel Sheet opened!")
        
    except:
        # Create a workbook
        wb = Workbook()

        # Change name of the initial worksheet
        ws = wb.active
        ws.title = "Owned Games"

        wb.create_sheet('Wishlist')

        wb.active = wb['Owned Games']

        # Make a new file 
        wb.save(filename = 'gamelist_wb.xlsx')
        print("Excel Sheet created!")
        print()

    finally:
        print("Now writing to Excel Sheet...")
        print()
        return wb

#looks at the workbook given, worksheet number, and inputs all data into the workbook
#next parameter is an integer: 0 = add to first worksheet, 1 = add to second worksheet
#data_all is of the form: [['gamename', [hours]], ...]
#the hours are floats
def addToWB(wb, wsnum, data_all):
    print('Adding game data to Excel sheet')

    #switches wsnum for the selected worksheet for the data to be input into
    if (wsnum == 0):
        wb.active = wb['Owned Games']
        ws = wb['Owned Games']
    elif (wsnum == 1):
        wb.active = wb['Wishlist']
        ws = wb['Wishlist']
    else:
        print('invalid worksheet')
        return -1

    #adjust first 3 columns to be wider for legibility
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    #list the column names in bold
    ws.cell(row = 1, column = 1, value = 'Game Name')
    ws['A1'].font = Font(bold=True)
    ws.cell(row = 1, column = 2, value = 'Main Story (Hours)')
    ws['B1'].font = Font(bold=True)
    ws.cell(row = 1, column = 3, value = 'Completionist (Hours)')
    ws['C1'].font = Font(bold=True)

    #adjust row height for column names
    ws.row_dimensions[1].height = 20

    #adjust all rows height to be 20 while inputting the data into the row
    for i in range(len(data_all)):
        ws.row_dimensions[i + 2].height = 20
        
        #add name of game
        ws.cell(row = i+2, column = 1, value = data_all[i][0])

        #add data from game if it exists
        #otherwise add 'no data'

        #data_all[index][0 = gamename, 1 = array of hours][which value of hours we want: 0 = main story, 1 = completionist]
        if str(type(data_all[i][1][0])) == '<class \'str\'>':
            #both the cells will contain the same data, since it doesn't matter
            ws.cell(row = i+2, column = 2, value = data_all[i][1][0])
            ws.cell(row = i+2, column = 3, value = data_all[i][1][0])
        else:
            ws.cell(row = i+2, column = 2, value = float(data_all[i][1][0]))
            ws.cell(row = i+2, column = 3, value = float(data_all[i][1][1]))

        ws.cell(row = i + 2, column = 5, value = None)

    #save work at the end
    wb.save(filename = 'gamelist_wb.xlsx')

    print("Done!")
    print()
    return

#colors the cells of the sheet according to:
#blue < 10 hours
#green < 20 hours
#purple < 35 hours
#yellow < 50 hours
#orange < 80 hours
#red >= 80 hours
def colorCoder(wb, wsnum):
    #switches wsnum for the selected worksheet for the data to be input into
    if (wsnum == 0):
        wb.active = wb['Owned Games']
        ws = wb['Owned Games']
    elif (wsnum == 1):
        wb.active = wb['Wishlist']
        ws = wb['Wishlist']
    else:
        print('invalid worksheet')
        return -1

    #check if there is data to be color coded
    #if so then perform color coding
    if ws.max_row <= 1:
        print('Nothing there!')
        return
    else:
        #color defaults FFFFFF
        blue = PatternFill(fill_type = 'solid', start_color = '87CEEB', end_color='87CEEB')
        green = PatternFill(fill_type = 'solid', start_color = '6be575', end_color='6be575')
        purple = PatternFill(fill_type = 'solid', start_color = 'B589ed', end_color='B589ed')
        yellow = PatternFill(fill_type = 'solid', start_color = 'E6ed89', end_color='E6ed89')
        orange = PatternFill(fill_type = 'solid', start_color = 'F1b521', end_color='F1b521')
        red = PatternFill(fill_type = 'solid', start_color = 'F36767', end_color='F36767')
        print('Color coding begins now!')
        print()

        #begin coloring the data starting from row 2, column 2
        #i.e. starting from cell 'B2'
        for i in range(2, ws.max_row + 1):
            for j in range(2,4):
                curr = ''
                if j == 2:
                    curr = 'B'
                elif j == 3:
                    curr = 'C'
                
                #append the number to the letter as a string
                curr = curr + str(i)

                #assign it to val which will then be checked if its a valid number or not
                val = ws[curr].value

                #if val is a number then it will be color coded
                #otherwise will be flagged for manual completion
                if not (str(type(val)) == '<class \'str\'>'):
                    #color code the cell
                    if val > 79:
                        ws[curr].fill = red
                    elif 80 > val and val > 50:
                        ws[curr].fill = orange
                    elif 51 > val and val > 35:
                        ws[curr].fill = yellow
                    elif 36 > val and val > 20:
                        ws[curr].fill = purple
                    elif 21 > val and val > 10:
                        ws[curr].fill = green
                    elif 11 > val:
                        ws[curr].fill = blue
                else:
                    print('NaN! - check ' + curr + ' manually')

        print()
        print('Done color coding!') 
        print()
        
        #save work at the end
        wb.save(filename = 'gamelist_wb.xlsx')  
                    
#gets rid of all the colors of 'Main Story' and 'Completionist' cells
def uncolorCoder(wb, wsnum):
    #switches wsnum for the selected worksheet for the data to be input into
    if (wsnum == 0):
        wb.active = wb['Owned Games']
        ws = wb['Owned Games']
    elif (wsnum == 1):
        wb.active = wb['Wishlist']
        ws = wb['Wishlist']
    else:
        print('invalid worksheet')
        return -1

    #check if there is data to be color coded
    #if so then perform color coding
    if ws.max_row <= 1:
        print('Nothing there!')
        return
    else:
        #begin de-coloring the data starting from row 2, column 2
        #i.e. starting from cell 'B2'
        for i in range(2, ws.max_row + 1):
            for j in range(2,4):
                curr = ''
                if j == 2:
                    curr = 'B'
                elif j == 3:
                    curr = 'C'
                
                #append the number to the letter as a string
                curr = curr + str(i)
                ws[curr].fill = PatternFill(fill_type = 'none')
        
        #save work at the end
        wb.save(filename = 'gamelist_wb.xlsx') 

#checks if a string contains only numerical characters
#i.e. 0-9    
#returns 0 or 1:
#       1 -> all numbers
#       0 -> not all numbers     
def onlyNum(string):
    valid = 1
    for i in range(len(string)):
        for j in range (10):
            if string[i] == str(j):
                valid = 1
                break
            else:
                valid = 0
        if not (valid == 0):
            break
    return valid

#reads a text file and returns an array of file contents, line by line
#returns an array with file contents if properly read
def fileReader(name):
    lines = []
    try:
        with open(name) as f:
            print('Reading the txt file!')
            for line in f:
                lines.append(line.strip())
            print('Done reading the txt file!')
            return lines
    
    except:
        print('ERROR READING TXT FILE: ' + name)
        return

#given a workbook, worksheet number, and name of txt file: will grab data from txt file and insert into first empty column
#within the worbkook at the given worksheet number
#the first line of the file will be the header of the column which will be bolded
#the rest of the lines will be input with no formatting
def addNewCol(wb, wsnum, name):

    print('Adding new Column of data from ' + name)
    print()

    #switches wsnum for the selected worksheet for the data to be input into
    if (wsnum == 0):
        wb.active = wb['Owned Games']
        ws = wb['Owned Games']
    elif (wsnum == 1):
        wb.active = wb['Wishlist']
        ws = wb['Wishlist']
    else:
        print('invalid worksheet')
        return -1

    #read data from file and add to array toAdd
    toAdd = fileReader(name)

    #for all items in toAdd, append them to the sheet at the 1st empty column
    for i in range (len(toAdd)):
        if i == 0:
            ws.cell(row = i + 1, column = ((ws.max_column) + 1), value = toAdd[i]).font = Font(bold=True)
        else:
            ws.cell(row = i + 1, column = (ws.max_column), value = toAdd[i])

    #save work at the end
    wb.save(filename = 'gamelist_wb.xlsx')

    print("Done! adding new column of data!")
    print()
    return

#given a workbook, worksheet number, and name of txt file: will grab data from txt file and insert into first empty row
#will append a new row of the given data
#data entered is an array with each entry being separate cell
def addNewRow(wb, wsnum, name):
    
    print('Adding new row(s) of data from ' + name)
    print()

    #switches wsnum for the selected worksheet for the data to be input into
    if (wsnum == 0):
        wb.active = wb['Owned Games']
        ws = wb['Owned Games']
    elif (wsnum == 1):
        wb.active = wb['Wishlist']
        ws = wb['Wishlist']
    else:
        print('invalid worksheet')
        return -1

    #read data from file and add to array toAdd
    #each line will be of the form: GAME NAME-PLATFORM+x
    #x indicates in progress
    #eg. Overwatch 2-PC+x
    toAdd = fileReader(name)
    allGames = []
    allPlatforms = []
    inProg = ''

    for i in range(len(toAdd)):
        #check if there is a space before the minus and grab the gamename
        if toAdd[i].find(' -') == -1:
            gamename = toAdd[i][:toAdd[i].find('-')]
        else:
            gamename = toAdd[i][:toAdd[i].find('-') - 1]

        if toAdd[i].find('+') == -1:
            platform = toAdd[i][toAdd[i].find('-') + 1:]
            inProg = ''
        else:
            platform = toAdd[i][toAdd[i].find('-') + 1:toAdd[i].find('+') - 1]
            inProg = 'x'
        

        allGames.append(gamename)
        allPlatforms.append(platform)

    #now we will perform the webscraping
    hours = webHandlr.beginWebScrape(allGames)
    print()

    #with all the data in hand, we shall input it into the workbook, at the specified worksheet
    for i in range(len(hours)):
        #gamename
        ws.cell(row = ws.max_row + 1, column = 1, value = hours[i][0])

        #hours
        if str(type(hours[i][1][0])) == '<class \'str\'>':
            ws.cell(row = ws.max_row, column = 2, value = hours[i][1][0])
            ws.cell(row = ws.max_row, column = 3, value = hours[i][1][0])
        else:
            ws.cell(row = ws.max_row, column = 2, value = float(hours[i][1][0]))
            ws.cell(row = ws.max_row, column = 3, value = float(hours[i][1][1]))

        #platform
        ws.cell(row = ws.max_row, column = 4, value = allPlatforms[i])

        #in progress
        if inProg == 'x':
            ws.cell(row = ws.max_row, column = 5, value = inProg)
        
    #we will colorcode it too
    colorCoder(wb, wsnum)

    #save work at the end
    wb.save(filename = 'gamelist_wb.xlsx')

    print("Done! adding new row of data!")
    return

#given a workbook, worksheet number, and sort number, will sort the data
#sort: 0 - gamename alphabetical order, 1 - Main story least to greatest, 2 - Completionist least to greatest, 3 - platform alphabetical
#0 = wb, 1 = wsnum, 2 = sort(1), 3 = sort2 
def sorter(*args):
    if (len(args) < 3):
        print('ERROR - NOT ENOUGH ARGUMENTS')
        return  
    elif (len(args) > 4):
        print('ERROR - TOO MANY ARGUMENTS')
        return
    else:
        #we sort the workbook given
        wb = args[0]
        wsnum = args[1]

        #switches wsnum for the selected worksheet for the data to be input into
        if (wsnum == 0):
            wb.active = wb['Owned Games']
            ws = wb['Owned Games']
        elif (wsnum == 1):
            wb.active = wb['Wishlist']
            ws = wb['Wishlist']
        else:
            print('invalid worksheet')
            return -1

        #in_progress will hold the in progress games which will be sorted separately than the rest, but added at the end
        #data_all will hold all the data which will be sorted
        data_all = []
        in_progress = []

        #grab data to be sorted
        for i in range(2, ws.max_row + 1):
            if ws['E' + str(i)].value == 'x':
                in_progress.append([ws['A' + str(i)].value, ws['B' + str(i)].value, ws['C' + str(i)].value, ws['D' + str(i)].value, ws['E' + str(i)].value])
            else:
                data_all.append([ws['A' + str(i)].value, ws['B' + str(i)].value, ws['C' + str(i)].value, ws['D' + str(i)].value])

        print('Beginning sort!')
        print()

        #if args[2] == 0, then alphanumerical sort
        #else perform args[2] sort then args[3] sort on the sublist to be sorted
        if args[2] == 0:
            #sort the data alphanumerically
            data_all = alphanumeriSort(data_all)

            #it doesn't matter if in_progress is empty or not in this case
            in_progress = alphanumeriSort(in_progress)
        else:
            #perform the first sort
            partitioned_data_all = []
            partitioned_in_progress = []

            if args[2] == 1:
                #Main story
                partitioned_data_all = hoursSort(data_all, 0)
                if not(in_progress == []):
                    partitioned_in_progress = hoursSort(in_progress, 0)
            elif args[2] == 2:
                #Completionist
                partitioned_data_all = hoursSort(data_all, 1)
                if not(in_progress == []):
                    partitioned_in_progress = hoursSort(in_progress, 0)
            elif args[2] == 3:
                #Platform
                partitioned_data_all = platformSort(data_all)
                if not(in_progress == []):
                    partitioned_in_progress = platformSort(in_progress)
            else:
                #invalid sort
                print('Invalid Sort ' + str(args[2]))
                return

            #clear data_all for inputting of the data
            data_all = []
            
            if not(in_progress == []):
                in_progress = []

            #perform the secondary sort if specified and isn't same as primary sort
            #otherwise perform alphanumeric sort by default
            if len(args) == 3 or args[3] == 0 or args[2] == args[3]:
                for i in range(len(partitioned_data_all)):
                    partitioned_data_all[i] = alphanumeriSort(partitioned_data_all[i])
                    for j in range(len(partitioned_data_all[i])):
                        data_all.append(partitioned_data_all[i][j])
                if not(partitioned_in_progress == []):
                    for i in range(len(partitioned_in_progress)):
                        partitioned_in_progress[i] = alphanumeriSort(partitioned_in_progress[i])
                        for j in range(len(partitioned_in_progress[i])):
                            in_progress.append(partitioned_in_progress[i][j])
            else:
                if not(partitioned_in_progress == []):
                    for i in partitioned_in_progress:
                        #check which secondary sort should be performed and have i contain the sorted list
                        if args[3] == 1:
                            i = hoursSort(i, 0)
                        elif args[3] == 2:
                            i = hoursSort(i, 1)
                        elif args[3] == 3:
                            i = platformSort(i)
                        else:
                            print('Invalid Sort ' + str(args[3]))
                            return

                        for j in i:
                            for k in j:
                                in_progress.append(k)

                for i in partitioned_data_all:
                    #check which secondary sort should be performed and have i contain the sorted list
                    if args[3] == 1:
                        i = hoursSort(i, 0)
                    elif args[3] == 2:
                        i = hoursSort(i, 1)
                    elif args[3] == 3:
                        i = platformSort(i)
                    else:
                        print('Invalid Sort ' + str(args[3]))
                        return

                    for j in i:
                        for k in j:
                            data_all.append(k)

        #data_all now holds the sorted data
        #data_all is of the form:
        #[['gamename', hours1, hours2, platform], ...]
        #print(data_all)

        #in_progress now holds the sorted data that is in progress
        #in_progress is of the form:
        #[['gamename', hours1, hours2, platform], ...]
        #print(in_progress)

        #uncolor the worksheet
        uncolorCoder(wb, wsnum)

        #update the worksheet
        #check if there are any in progress and input those first. afterwards input rest of data
        #otherwise just put data as is
        if len(in_progress) > 0:
            for i in range(len(in_progress)):
                #input gamename
                ws.cell(row = i + 2, column = 1, value = in_progress[i][0])

                #if the hours for main game/completionist are of string, then put in as a string
                #otherwise put in as a float
                if str(type(in_progress[i][1])) == '<class \'str\'>':
                    ws.cell(row = i + 2, column = 2, value = in_progress[i][1])
                else:
                    ws.cell(row = i + 2, column = 2, value = float(in_progress[i][1]))
            
                if str(type(in_progress[i][2])) == '<class \'str\'>':
                    ws.cell(row = i + 2, column = 3, value = in_progress[i][2])
                else:
                    ws.cell(row = i + 2, column = 3, value = float(in_progress[i][2]))
                
                #input platform
                ws.cell(row = i + 2, column = 4, value = in_progress[i][3])

                #in progress
                ws.cell(row = i + 2, column = 5, value = in_progress[i][4])

            for i in range(len(data_all)):
                #input gamename
                ws.cell(row = i + 2 + len(in_progress), column = 1, value = data_all[i][0])

                #if the hours for main game/completionist are of string, then put in as a string
                #otherwise put in as a float
                if str(type(data_all[i][1])) == '<class \'str\'>':
                    ws.cell(row = i + 2 + len(in_progress), column = 2, value = data_all[i][1])
                else:
                    ws.cell(row = i + 2 + len(in_progress), column = 2, value = float(data_all[i][1]))
            
                if str(type(data_all[i][2])) == '<class \'str\'>':
                    ws.cell(row = i + 2 + len(in_progress), column = 3, value = data_all[i][2])
                else:
                    ws.cell(row = i + 2 + len(in_progress), column = 3, value = float(data_all[i][2]))
                
                #input platform
                ws.cell(row = i + 2 + len(in_progress), column = 4, value = data_all[i][3])

                ws.cell(row = i + 2 + len(in_progress), column = 5, value = '')
        else:
            for i in range(len(data_all)):
                #input gamename
                ws.cell(row = i + 2, column = 1, value = data_all[i][0])

                #if the hours for main game/completionist are of string, then put in as a string
                #otherwise put in as a float
                if str(type(data_all[i][1])) == '<class \'str\'>':
                    ws.cell(row = i + 2, column = 2, value = data_all[i][1])
                else:
                    ws.cell(row = i + 2, column = 2, value = float(data_all[i][1]))
            
                if str(type(data_all[i][2])) == '<class \'str\'>':
                    ws.cell(row = i + 2, column = 3, value = data_all[i][2])
                else:
                    ws.cell(row = i + 2, column = 3, value = float(data_all[i][2]))
                
                #input platform
                ws.cell(row = i + 2, column = 4, value = data_all[i][3])

                ws.cell(row = i + 2 + len(in_progress), column = 5, value = '')

        #color code the worksheet
        colorCoder(wb, wsnum)

        #save the worksheet
        wb.save(filename = 'gamelist_wb.xlsx')

        print('Done sorting!')
        print()

#sorts the given list's elements first element alphanumerically: special characters, 0-9, a-z, A-Z
#eg. [['Overwatch 2', ...], ['Final Fantasy', ...], ...]
#returns the given list sorted
def alphanumeriSort(list):
    for i in range(1, len(list)):
        #insertion sort
        j = i - 1
        print(list[i])
        while not (j == -1):   
            if list[i][0].casefold() <= list[j][0].casefold():
                list.insert(j, list.pop(i))
                i -= 1
            j -= 1
    return list

#sorts the given list's elements by Main story hours
#eg. [['Overwatch 2', ...], ['Final Fantasy', ...], ...]
#returns the given list sorted by Main story hours from least to greatest
#hoursType = 0 -> Main story
#hoursType = 1 -> Completionist
def hoursSort(list, hoursType):
    partitionedList = [[list[0]]]
    if not((hoursType == 0) or (hoursType == 1)):
        print('Invalid hoursSort hoursType')
        return [list]

    for i in range(1, len(list)):
        #insert the game into its' respective grouping for the platform
        for j in range(len(partitionedList)):
            result = 0

            if hoursType == 0:
                #will do for main story
                result = list[i][1] == partitionedList[j][0][1]
            elif hoursType == 1:
                #will do for completionist
                result = list[i][2] == partitionedList[j][0][2]

            if result == 1:
                partitionedList[j].append(list[i])
                break

            if j == (len(partitionedList) - 1):
                partitionedList.append([list[i]])

    #now we have them all grouped up individually based off of hours that are the same
    #however there may be non-numerical values that need to be not sorted with the numerical values
    #nonNum tells how many non-numerical values there are
    nonNum = 0

    #increment nonNum, and move the non-numerically grouped hours together to avoid being sorted
    for i in range(len(partitionedList)):
        if hoursType == 0 and str(type(partitionedList[i][0][1])) == '<class \'str\'>':
            partitionedList.insert(0, partitionedList.pop(i))
            nonNum += 1
        elif hoursType == 1 and str(type(partitionedList[i][0][2])) == '<class \'str\'>':
            partitionedList.insert(0, partitionedList.pop(i))
            nonNum += 1

    #insertion sort on the numerical hours that are listed
    for i in range(1 + nonNum, len(partitionedList)):
        j = i - 1
        while not (j == nonNum - 1):
            result = 0

            if hoursType == 0:
                result = partitionedList[j][0][1] - partitionedList[i][0][1]
            elif hoursType == 1:
                result = partitionedList[j][0][2] - partitionedList[i][0][2]

            if result > 0:
                partitionedList.insert(j, partitionedList.pop(i))
                i -= 1
            j -= 1

    return partitionedList

#sorts the given list's elements by platform
#eg. [['Overwatch 2', ...], ['Final Fantasy', ...], ...]
#returns the given list sorted by platform
def platformSort(list):
    #partitionedList will contain separate arrays that hold each of the platforms (PC, PS4, etc.)
    partitionedList = [[list[0]]]
    for i in range(1, len(list)):
        #insert the game into its' respective grouping for the platform
        for j in range(len(partitionedList)):
            if list[i][3] == partitionedList[j][0][3]:
                partitionedList[j].append(list[i])
                break
            if j == (len(partitionedList) - 1):
                partitionedList.append([list[i]])

    #partitionedList now contains each of the games separated into categories based off of the platform they are    
    #print(partitionedList)
    return partitionedList

###################################################################
#                   Commands Section
#sorter(wbChecker(), 0, 1, 2)
